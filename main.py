from selenium import webdriver
import openpyxl
import time
from selenium.webdriver.common.action_chains import ActionChains
browser=webdriver.Firefox()
browser.get('http://www.farmsoft.com.cn/lord/')
def login_system():    #输入用户名密码登录系统
    farm_element=browser.find_element_by_css_selector('.textbox-icon')
    farm_element.click()
    time.sleep(1)
    nkzj_element=browser.find_element_by_css_selector('#_easyui_tree_129 > span:nth-child(1)')
    nkzj_element.click()
    time.sleep(1)
    nc_element=browser.find_element_by_css_selector('#_easyui_tree_143 > span:nth-child(4)')
    nc_element.click()
    user_element=browser.find_element_by_css_selector('.top > div:nth-child(2) > div:nth-child(1) > span:nth-child(2) > input:nth-child(1)')
    user_element.send_keys('liankeyu')#输入用户名
    psw_element=browser.find_element_by_css_selector('.top > div:nth-child(3) > div:nth-child(1) > span:nth-child(2) > input:nth-child(1)')
    psw_element.send_keys('xxxxxx')#输入密码
    login_element=browser.find_element_by_css_selector('#login')
    login_element.click()
def turn2Dgzzjl():#进入稻谷种植记录
    time.sleep(1)
    xxjlgl_element=browser.find_element_by_css_selector('#menu-accordion > div:nth-child(2) > div:nth-child(1) > div:nth-child(1)')
    xxjlgl_element.click()
    dgzzjl_element=browser.find_element_by_css_selector('#_easyui_tree_7 > span:nth-child(3) > div:nth-child(1)')
    dgzzjl_element.click()
#进入填写表单界面，需要打开新的网页
def turn2Newpage():
    js='window.open("http://www.farmsoft.com.cn/lord/BizForm/../Form/DimEdit?formId=19&parentId=");'
    browser.execute_script(js)
    browser.switch_to_window(browser.window_handles[-1])
def FillTheForm():
    #填写表单
    time.sleep(2)
    browser.find_element_by_css_selector('#datagrid-row-r1-2-'+I+' > td:nth-child(2) > div:nth-child(1) > table:nth-child(1) > tbody:nth-child(1) > tr:nth-child(1) > td:nth-child(1) > span:nth-child(2) > span:nth-child(1) > a:nth-child(1)').click()
    time.sleep(1)
    browser.find_element_by_css_selector('.tree-hit').click()
    time.sleep(3)
    #根据人名排序选择点击人名
    zzh=browser.find_element_by_css_selector('div.combo-p:nth-child(9) > div:nth-child(1) > ul:nth-child(1) > li:nth-child(1) > ul:nth-child(2) > li:nth-child('+zzhf+')')
    zzh.click()#根据人名排列顺序index（从1到最后）
    time.sleep(1)
    #填入品种
    pz=browser.find_element_by_css_selector('#datagrid-row-r1-2-'+I+' > td:nth-child(4) > div:nth-child(1) > table:nth-child(1) > tbody:nth-child(1) > tr:nth-child(1) > td:nth-child(1) > span:nth-child(2) > input:nth-child(2)')
    pz.send_keys(str(pzf))
    #填入日期
    bzrq=browser.find_element_by_css_selector('.datebox > input:nth-child(2)')
    bzrq.send_keys(str(bzrqf))
    #填入种植面积
    mianji=browser.find_element_by_css_selector('input.validatebox-invalid:nth-child(1)')
    mianji.send_keys(str(mianjif))
    #填入计量单位
    jiliangdanwei=browser.find_element_by_css_selector('#datagrid-row-r1-2-'+I+' > td:nth-child(8) > div:nth-child(1) > table:nth-child(1) > tbody:nth-child(1) > tr:nth-child(1) > td:nth-child(1) > span:nth-child(2) > span:nth-child(1) > a:nth-child(1)')
    jiliangdanwei.click()
    time.sleep(1)
    mushu=browser.find_element_by_css_selector('#_easyui_combobox_i'+str(int(I)+1)+'_0')
    mushu.click()
    #填入种子来源
    zhongzilaiyuan=browser.find_element_by_css_selector('#datagrid-row-r1-2-'+I+' > td:nth-child(10) > div:nth-child(1) > table:nth-child(1) > tbody:nth-child(1) > tr:nth-child(1) > td:nth-child(1) > span:nth-child(2) > span:nth-child(1) > a:nth-child(1)')
    zhongzilaiyuan.click()
    time.sleep(2)
    #种子来源选择倒数第二项
    zhongzigongsi=browser.find_element_by_css_selector('div.panel:nth-child(13) > div:nth-child(1) > ul:nth-child(1) > li:nth-child(23)')#里面的2就是第二项
    zhongzigongsi.click()
    time.sleep(2)
    if K==1:
        ActionChains(browser).move_by_offset(33, 150.5).click().perform()
    else:
        ActionChains(browser).move_by_offset(0, 32).click().perform()
def Saveandnew():
    #保存表单并继续
    Save=browser.find_element_by_css_selector('#btnSaveAndNew')
    #Save.click()
wb=openpyxl.load_workbook('project1.xlsx')
sheet=wb.get_sheet_by_name('稻谷种植记录')
#取得最大行数
login_system()#登录系统
turn2Dgzzjl()#打开稻谷种植记录
turn2Newpage()
#找到数据并赋值
max_Hang=sheet.max_row
K=1
for i in range(2,max_Hang+1):
    zzhf=sheet.cell(row=i,column=1).value
    pzf=sheet.cell(row=i,column=3).value
    bzrqf=sheet.cell(row=i,column=4).value
    mianjif=sheet.cell(row=i,column=5).value
    zzlyf=sheet.cell(row=i,column=7).value
    if int(zzhf)>=10:
        I=str(int(zzhf)-1)
    else:
        I=str(int(zzhf)%10-1)
    FillTheForm()
    K+=1
    if i%10==0:
        print('该按保存并继续了')
        K=1
        time.sleep(2)
